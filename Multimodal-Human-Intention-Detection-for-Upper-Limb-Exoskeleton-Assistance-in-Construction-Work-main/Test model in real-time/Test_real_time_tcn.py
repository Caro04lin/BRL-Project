# Should be run after all data have been collected
import os
import re
import openpyxl
from statistics import mean
import sys

# Define folder paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TXT_FILE = input("Enter the name of the TXT file (without extension) in TCN folder: ").strip()
TXT_PATH = os.path.join(BASE_DIR, "TCN", TXT_FILE + ".txt")
EXCEL_PATH = os.path.join(BASE_DIR, "Results_AI.xlsm")

# Verify file existence
if not os.path.exists(TXT_PATH):
    raise FileNotFoundError(f"TXT file not found at {TXT_PATH}")
if not os.path.exists(EXCEL_PATH):
    raise FileNotFoundError(f"Excel file not found at {EXCEL_PATH}")

# Load Excel workbook
wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
print(f"Sheets available: {wb.sheetnames}")
sheet_name = input("Enter the sheet name to record the results: ").strip()

if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

ws = wb[sheet_name]

# Read the log file
with open(TXT_PATH, "r", encoding="utf-8") as f:
    lines = [line.strip() for line in f.readlines() if line.strip()]

# Regex pattern: extract sample name, action, tool, and time
pattern = r"(Sample_\d+)\s*:\s*([\w_]+)\s*\|\s*Tool\s*:\s*([\w_]+)\s*at\s*([\d.]+)"
parsed_data = []

for line in lines:
    match = re.match(pattern, line)
    if match:
        parsed_data.append(match.groups())

if not parsed_data:
    raise ValueError("No valid prediction lines found in the TXT file.")

# --- Write data to Excel ---
start_row = 3
test_counter = 0  # Count the number of test blocks

for i, (sample_name, action, tool, pred_time) in enumerate(parsed_data, start=start_row):
    # Increment test number only when encountering 'Sample_1'
    if sample_name == "Sample_1":
        test_counter += 1
        ws[f"I{i}"] = test_counter  # Write test number
    else:
        ws[f"I{i}"] = None  # Leave empty if not Sample_1

    # Write number of sample in column J
    ws[f"J{i}"] = i - start_row + 1  # Incremental number of sample

    # Write the rest of the data
    ws[f"K{i}"] = action           # Action name
    ws[f"L{i}"] = tool             # Tool used
    ws[f"M{i}"] = float(pred_time) # Prediction time

# --- Compute accuracy ---
expected_action = ws["A2"].value
predicted_actions = [ws[f"K{row}"].value for row in range(start_row, start_row + len(parsed_data))]
correct = sum(1 for a in predicted_actions if a == expected_action)
accuracy = (correct / len(predicted_actions)) * 100 if predicted_actions else 0
ws["N3"] = round(accuracy, 2)

# --- Group prediction times by test block ---
sample_blocks = {}
current_test = 0
for i in range(start_row, start_row + len(parsed_data)):
    test_num = ws[f"I{i}"].value
    pred_time = ws[f"M{i}"].value

    # If a new test starts (Sample_1 line)
    if test_num is not None:
        current_test = test_num
        sample_blocks[current_test] = []

    # Add prediction time to the current block
    if current_test != 0:
        sample_blocks[current_test].append(pred_time)

# --- Compute mean time difference per test ---
mean_diffs = []
for block_times in sample_blocks.values():
    if len(block_times) > 1:
        diffs = [block_times[j + 1] - block_times[j] for j in range(len(block_times) - 1)]
        mean_diffs.append(mean(diffs))

# Write overall mean of all block averages to O3
ws["O3"] = round(mean(mean_diffs), 3) if mean_diffs else "N/A"

# --- Compute mean of all Sample_1 prediction times ---
first_samples = [times[0] for times in sample_blocks.values() if times]
ws["O5"] = round(mean(first_samples), 3) if first_samples else "N/A"

# --- Write total number of samples ---
ws["O8"] = len(parsed_data)

# --- Debug information ---
print(f"✅ Detected {len(sample_blocks)} test blocks based on 'Sample_1'.")
print(f"✅ Total samples: {len(parsed_data)}")
for test, times in sample_blocks.items():
    print(f"Test {test}: {len(times)} samples, first={times[0]}, last={times[-1]}")

# --- Safe saving ---
try:
    wb.save(EXCEL_PATH)
    print(f"✅ Results successfully written to '{EXCEL_PATH}'")
except PermissionError:
    print("⚠️ Please close the Excel file before running the script again.")
    sys.exit(1)


